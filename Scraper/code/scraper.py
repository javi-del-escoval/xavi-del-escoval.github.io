from requests import get
from bs4 import BeautifulSoup
from pandas import DataFrame, read_excel
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from playwright.sync_api import sync_playwright
from playwright._impl._errors import TimeoutError
from tkinter_unblur import Tk
from tkinter import StringVar
from tkinter.messagebox import showinfo, showerror, askokcancel, ERROR
from traceback import format_exc
from json import load
from re import search
from datetime import date, datetime, timedelta
from time import sleep
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from os import path, startfile
from reset_config import reset_config

'''
┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
┃				UTILS				┃
┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛'''
def get_data(filename):
	output = []
	try:
		with open(filename, 'r', encoding="utf-8") as file:
			for line in file:
				output.append(line.strip())
	except Exception as e:
		print(f'[{filename}] Error: {e}\n{format_exc()}')
	return output

def export_to_excel(filename, df):
	try:
		df['se toma'] = False
		df = df.drop_duplicates()
		df.to_excel(filename, index=False)
		wb = load_workbook(filename=filename)
		ws = wb.active
		ws.column_dimensions['A'].width = 75
		ws.column_dimensions['B'].width = 20
		ws.column_dimensions['C'].width = 20
		ws.column_dimensions['D'].width = 75
		ws.column_dimensions['E'].width = 10
		for column in [ws['A'], ws['B'], ws['C'], ws['D'], ws['E']]:
			for row in column:
				row.alignment = Alignment(wrap_text=True)
		wb.save(filename)
	except Exception as e:
		if not askokcancel('Web Scraper - ECIT', f'[Export] Error: {e}\n{format_exc()}', icon=ERROR):
			return False, False, False

'''
┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
┃			CONFIGURACION			┃
┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛'''
_DATA = {}
try:
	with open('config.json', 'r', encoding='utf-8') as file:
		_DATA = load(file)
except FileNotFoundError:
	reset_config()
	try:
		with open('config.json', 'r', encoding='utf-8') as file:
			_DATA = load(file)
	except Exception as e:
		showerror('Web Scraper - ECIT', f'[Load Config][After Reset] Error: {type(e)}\n{e}\n{format_exc()}')
except Exception as e:
	showerror('Web Scraper - ECIT', f'[Load Config] Error: {type(e)}\n{e}\n{format_exc()}')

_DEBUG = False

'''
┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
┃		  MERCADO PUBLICO			┃
┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛'''
def mercado_publico(ticket=_DATA['ticket'], keywords=_DATA['keywords'], dias = 7, hasta = None):
	results = []
	excluded = 0
	count = 0

	hoy = date.today()
	if hasta:
		hoy = hasta
	fechas = [(hoy-timedelta(days=dias-i)).strftime('%d%m%Y') for i in range(dias+1)]
	for fecha in fechas:
		try:
			url = f'https://api.mercadopublico.cl/servicios/v1/publico/licitaciones.json?fecha={fecha}&estado=publicada&ticket={ticket}'
			response = get(url)
			if response.status_code == 200:
				data = response.json()
				for entry in data['Listado']:
					if any(keyword in entry['Nombre'] for keyword in keywords):
						code = entry['CodigoExterno']
						url = f'https://api.mercadopublico.cl/servicios/v1/publico/licitaciones.json?codigo={code}&ticket={ticket}'
						response = get(url)
						if response.status_code == 200:
							details = response.json()['Listado']
							inicio = ''
							descripcion = ''
							if isinstance(details, dict):
								if isinstance(details['Fechas'], dict):
									tmp = details['Fechas']['FechaInicio']
									inicio = datetime(tmp[:4], tmp[5:7], tmp[8:10], tmp[11:13], tmp[14:16], tmp[17:19])
								else:
									if not askokcancel('Web Scraper - ECIT', f'TypeError\n{type(details['Fechas'])}\n{details['Fechas']}\n\n', icon=ERROR):
										return False, False, False
								if isinstance(details['Items'], dict) and isinstance(details['Items']['Listado'], list) and isinstance(details['Items']['Listado'][0], dict):
									descripcion = details['Items']['Listado'][0]['Descripcion']
								else:
									if not askokcancel('Web Scraper - ECIT', f'TypeError\n{type(details['Items'])}\n{details['Items']}\n\n', icon=ERROR):
										return False, False, False
							else:
								if not askokcancel('Web Scraper - ECIT', f'TypeError\n{type(details)}\n{details}\n\n', icon=ERROR):
									return False, False, False
							tmp = entry['FechaCierre']
							fin = datetime(tmp[:4], tmp[5:7], tmp[8:10], tmp[11:13], tmp[14:16], tmp[17:19])
							if not any(exclusion in entry['Nombre'] for exclusion in _DATA['exclusions']) and (fin - datetime.now()).days > 0:
								results.append(
								{
									'titulo'		:	entry['Nombre'],
									'fecha-inicio'	:	inicio,
									'fecha-fin'		:	fin,
									'descripcion'	:	descripcion,
									'plataforma'	:	'Mercado Publico',
									'anunciante'	:	details['NombreOrganismo']+' '+details['NombreUnidad']
								})
								count += 1
							else:
								excluded +=1
								if _DEBUG:
									print(f'[Mercado Publico] Excluido')
					else:
						excluded += 1
						if _DEBUG:
							print('[Mercado Publico] Excluido')
		except Exception as e:
			if not askokcancel('Web Scraper - ECIT', f'[Mercado Publico] Error: {e}\n{format_exc()}', icon=ERROR):
				return False, False, False
		sleep(1) # para no saturar el server
	return results, count, excluded


'''
┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
┃				CORFO				┃
┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛'''

def corfo(keyword):
	URL = f"https://www.corfo.gob.cl/sites/cpp/programasyconvocatorias/?search={keyword}"

	results = []
	excluded = 0
	count = 0

	try:
		with sync_playwright() as p:
			browser = p.chromium.launch(headless=False)
			page = browser.new_page(
				user_agent=	"Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
							"AppleWebKit/537.36 (KHTML, like Gecko) "
							"Chrome/120.0.0.0 Safari/537.36",
				locale="es-CL",
				timezone_id="America/Santiago",
				viewport={"width": 1366, "height": 768}
			)


			try:
				page.goto(URL, timeout=60000)

				# Espera explícita de carga de resultados
				page.wait_for_selector('div.contenido-caja_prog', timeout=5000)

				boxes = page.locator('div.contenido-caja_prog').all()
				for box in boxes:
					try:
						data = box.inner_text().split('\n')
						inicio = '' if len(data) < 6 else datetime.strptime(data[2],'%d/%m/%Y').date()
						fin = '' if len(data) < 8 else datetime.strptime(data[4],'%d/%m/%Y').date()
						if not any(exclusion in data[0] for exclusion in _DATA['exclusions']):
							desc = data[1] if len(data) < 6 else data[5] if len(data) < 9 else data[8]
							if type(fin) == type(date.today()):
								if (fin - date.today()).days > 0:
									results.append(
									{
										'titulo'		:	data[0],
										'fecha-inicio'	:	inicio,
										'fecha-fin'		:	fin,
										'descripcion'	:	desc if desc.strip() != '' else f'https://www.corfo.gob.cl/sites/cpp/programasyconvocatorias/?search={data[0]}',
										'plataforma'	:	'CORFO',
										'anunciante'	:	''
									})
									count += 1
							else:
								results.append(
								{
									'titulo'		:	data[0],
									'fecha-inicio'	:	inicio,
									'fecha-fin'		:	fin,
									'descripcion'	:	desc if desc.strip() != '' else f'https://www.corfo.gob.cl/sites/cpp/programasyconvocatorias/?search={data[0]}',
									'plataforma'	:	'CORFO',
									'anunciante'	:	''
								})
								count += 1
						else:
							excluded += 1
							if _DEBUG:
								print(f'[CORFO][{keyword}] Excluido')
					except Exception as e:
						if not askokcancel('Web Scraper - ECIT', f'[CORFO][{keyword}]{e}', icon=ERROR):
							return False, False, False
			except TimeoutError:
				print(f'[CORFO][{keyword}] No Results')
			except Exception as e:
				if not askokcancel('Web Scraper - ECIT', f'[CORFO][{keyword}]{e}', icon=ERROR):
					return False, False, False
			finally:
				browser.close()
	except Exception as e:
		if not askokcancel('Web Scraper - ECIT', f'[CORFO][{keyword}]{e}', icon=ERROR):
			return False, False, False
	finally:
		return results, count, excluded


'''
┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
┃				BID					┃
┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛'''

def bid():
	results = []
	excluded = 0
	count = 0

	driver = webdriver.Chrome()

	target_url = 'https://beo-procurement.iadb.org/es/home'

	try:
		driver.get(target_url)
		sleep(3) # wait for the page to load
		page_source = driver.page_source

		soup = BeautifulSoup(page_source, 'html.parser')

		result_boxes = soup.find_all('tr', class_='master-row')

		for box in result_boxes:
			try:
				tds = box.find_all('td')
				cierre = box.find('td', class_='oei-date')

				tmp_div = box.find('div', class_='col-sm-12 col-md-6')
				txt_apertura = ''
				if tmp_div != None:
					txt_apertura = tmp_div.inner_text()
				match = search(r'\d\d-(.*)\d\d\d\d<',txt_apertura)
				inicio = ''
				if match:
					inicio = match.group()[:-1]
					inicio.strip()

				meses = { "Ene": "Jan", "Feb": "Feb", "Mar": "Mar", "Abr": "Apr", "May": "May", "Jun": "Jun", "Jul": "Jul", "Ago": "Aug", "Sep": "Sep", "Oct": "Oct", "Nov": "Nov", "Dic": "Dec" }
				fecha_str = cierre.text[:-24]
				for es, en in meses.items():
					fecha_str = fecha_str.replace(f"-{es}-", f"-{en}-")
				fin = datetime.strptime(fecha_str, '%d-%b-%Y %I:%M %p')

				desc = 'Error al extraer link'
				link = tds[1].find('a')
				if link and link['href']:
					desc = link['href']

				if not any(exclusion in tds[2].text for exclusion in _DATA['exclusions']) and (fin - datetime.now()).days > 0:
					results.append({
									'titulo'		:	tds[2].text,
									'fecha-inicio'	:	inicio,
									'fecha-fin'		:	fin,
									'descripcion'	:	desc,
									'plataforma'	:	'BID',
									'anunciante'	:	''
								})
					count += 1
				else:
					excluded += 1
					if _DEBUG:
						if not askokcancel('Web Scraper - ECIT', f'[BID] Excluido', icon=ERROR):
							return False, False, False
			except Exception as e:
				if not askokcancel('Web Scraper - ECIT', f'[BID] Error: {e}\n{format_exc()}', icon=ERROR):
					return False, False, False
	except Exception as e:
		if not askokcancel('Web Scraper - ECIT', f'[BID] Error: {e}\n{format_exc()}', icon=ERROR):
			return False, False, False
	finally:
		driver.quit()
		return results, count, excluded


'''
┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
┃			  CODESSER				┃
┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛'''

def codesser(search_query):
	results = []
	excluded = 0
	closed = 0
	count = 0

	driver = webdriver.Chrome()

	# The target URL with the search bar
	target_url = "https://codesser.cl/licitaciones/"

	try:
		driver.get(target_url)
		sleep(1) # wait for the page to load
		
		#WebDriverWait(driver, 10).until(EC.presence_of_element_located())
		div_element = driver.find_element(By.CLASS_NAME, "input-group")
		search_bar = div_element.find_element(By.TAG_NAME, "input")

		search_bar.send_keys(search_query)
		search_bar.send_keys(Keys.RETURN)

		#sleep(2) # wait for the page to load

		page_source = driver.page_source

		soup = BeautifulSoup(page_source, 'html.parser')

		# Extract Data
		rows = soup.find_all('tr')
		for row in rows:
			try:
				data = row.find_all('td')
				fin = ''
				if len(data) > 9:
					tmp = data[6].text.strip()
					fin = datetime.strptime(tmp, '%d-%b-%y')
					if data[7].text != '':
						match_obj = search(r'\d+:\d+', data[7].text)
						if match_obj:
							tmp += ' ' + match_obj.group()
							tmp = tmp.strip()
							fin = datetime.strptime(tmp, '%d-%b-%y %H:%M')

					if not any(exclusion in data[3].text for exclusion in _DATA['exclusions']) and (fin - datetime.now()).days > 0: # and any(region in data[4].text for region in ['Nacional', 'Región Metropolitana de Santiago'])
						results.append({
									'titulo'		:	data[3].text,
									'fecha-inicio'	:	datetime.strptime(data[5].text, '%d-%b-%y').date(),
									'fecha-fin'		:	fin,
									'descripcion'	:	data[9].a.href,
									'plataforma'	:	'CODESSER',
									'anunciante'	:	data[8].text
								})
						count += 1
				else:
					excluded += 1
					if type(fin) != type('') and (fin - datetime.now()).days > 0:
						closed += 1
					if _DEBUG:
						print(f'[CODESSER] Excluido')
			except Exception as e:
				if not askokcancel('Web Scraper - ECIT', f'[CODESSER][{search_query}] Error: {e}\n{format_exc()}', icon=ERROR):
					return False, False, False
				print(f'{format_exc()}')
	except Exception as e:
		if not askokcancel('Web Scraper - ECIT', f'[CODESSER][{search_query}] Error: {e}\n{format_exc()}', icon=ERROR):
			return False, False, False
		print(f'{format_exc()}')
	finally:
		driver.quit()
		if _DEBUG:
			showinfo('Web Scraper - ECIT', f'[CODESSER][{search_query}] {closed} resultados estaban cerrados')
		return results, count, excluded


'''
┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
┃				ANID				┃
┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛'''

def anid():
	results = []
	excluded = 0
	count = 0

	url = "https://anid.cl/concursos/"
	headers = {
		"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
					  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
		"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
		"Accept-Language": "es-CL,es;q=0.9",
		"Referer": "https://anid.cl/concursos/"
	}
	response = get(url, headers=headers, timeout=5)

	soup = BeautifulSoup(response.text, 'html.parser')

	try:
		boxes = soup.find_all('div', class_='jet-listing-grid__item')
		for i in range(0, len(boxes), 2):
			container_box = boxes[i]
			try:
				box = container_box.find('div', class_='e-con-inner')
				link = container_box.find('a', class_='jet-engine-listing-overlay-link')
				title = ''
				title_h3 = box.find('h3')
				if not title_h3 is None:
					title = title_h3.decode_contents()

				content = box.decode_contents()

				meses = { "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6, "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12 }
				inicio = ''
				start_match = search(r'Inicio: (.*)<?', content)
				if start_match:
					fecha_inicio = start_match.group()
					fecha_inicio = fecha_inicio.replace('</div>','')
					fecha_inicio = fecha_inicio.replace('Inicio: ','')
					fecha_inicio = fecha_inicio.strip()
					try:
						dia, resto = fecha_inicio.split(" de ", 1)
						mes_str, anio = resto.replace(",", "").split()
						inicio = datetime(int(anio),meses[mes_str.lower()],int(dia))
					except Exception as e:
						if not askokcancel('Web Scraper - ECIT', f'[ANID] Error: {e}\n{format_exc()}\n{fecha_inicio}', icon=ERROR):
							return False, False, False
				
				fin = ''
				end_match = search(r'Cierre: (.*)<?', content)
				if end_match:
					fecha_fin = end_match.group()
					fecha_fin = fecha_fin.replace('</div>','')
					fecha_fin = fecha_fin.replace('Cierre: ','')
					fecha_fin = fecha_fin.strip()
					try:
						dia, resto = fecha_fin.split(" de ", 1)
						mes_str, anio, dash, time = resto.replace(",", "").split()
						hora, minutos = time.split(":")
						fin = datetime(int(anio),meses[mes_str.lower()],int(dia), int(hora), int(minutos))
					except Exception as e:
						if not askokcancel('Web Scraper - ECIT', f'[ANID] Error: {e}\n{format_exc()}\n{fecha_fin}', icon=ERROR):
							return False, False, False

				desc = 'Error al extraer link'
				if link and link['href']:
					desc = link['href']
				if title != '' and fecha_fin != '' and not any(exclusion in title for exclusion in _DATA['exclusions']):
					if fin != '':
						if (fin - datetime.now()).days > 0:
							results.append(
							{
								'titulo'		:	title,
								'fecha-inicio'	:	inicio,
								'fecha-fin'		:	fin,
								'descripcion'	:	desc,
								'plataforma'	:	'ANID',
								'anunciante'	:	''
							})
							count += 1
						else:
							excluded += 1
							if _DEBUG:
								print(f'[ANID] Excluido')
					else:
						results.append(
						{
							'titulo'		:	title,
							'fecha-inicio'	:	inicio,
							'fecha-fin'		:	fin,
							'descripcion'	:	desc,
							'plataforma'	:	'ANID',
							'anunciante'	:	''
						})
						count += 1
				else:
					excluded += 1
					if _DEBUG:
						print(f'[ANID] Excluido')
			except Exception as e:
				if not askokcancel('Web Scraper - ECIT', f'[ANID] Error: {e}\n{format_exc()}', icon=ERROR):
					return False, False, False
	except Exception as e:
		if not askokcancel('Web Scraper - ECIT', f'[ANID] Error: {e}\n{format_exc()}', icon=ERROR):
			return False, False, False
	finally:
		return results, count, excluded

'''
┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
┃			  SERCOTEC				┃
┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛'''
def sercotec():
	results = []
	excluded = 0
	count = 0

	target_url = 'https://www.sercotec.cl/'

	try:
		with sync_playwright() as p:
			browser = p.chromium.launch(headless=False)
			page = browser.new_page()

			page.goto(target_url, wait_until="domcontentloaded", timeout=5000)
			page.wait_for_load_state("networkidle")

			frame = page.frame_locator("iframe.otrocalendario")
			boxes = frame.locator('#concursos').all()
			for box in boxes:
				try:
					concursos = box.locator('li').all()
					for concurso in concursos:
						try:
							tmp = concurso.locator('.inicio-btn').inner_text().lower().replace('inicio', '').strip()
							if len(tmp) > 10:
								inicio = datetime.strptime(tmp, '%d-%m-%Y %H:%M')
							else:
								inicio = datetime.strptime(tmp, '%d-%m-%Y')
							obj = concurso.locator('a')
							if not any(exclusion in obj.inner_text() for exclusion in _DATA['exclusions']): # and (concurso.locator('.region').inner_text() == 'REGIÓN METROPOLITANA' or concurso.locator('.instrumento').inner_text() == 'REGIÓN METROPOLITANA' or concurso.locator('.region').inner_text() == 'Todo Chile' or concurso.locator('.instrumento').inner_text() == 'Todo Chile')
								results.append({
									'titulo'		:	obj.inner_text(),
									'fecha-inicio'	:	inicio,
									'fecha-fin'		:	'',
									'descripcion'	:	obj.get_attribute('href'),
									'plataforma'	:	'SERCOTEC',
									'anunciante'	:	''
								})
								count += 1
							else:
								excluded += 1
								if _DEBUG:
							 		print(f'[SERCOTEC] Excluido\n{obj.inner_text()}')
						except TimeoutError:
							print('[SERCOTEC] No results')
						except Exception as e:
							if not askokcancel('Web Scraper - ECIT', f'[SERCOTEC] Error: {type(e)}\n{e}', icon=ERROR):
								return False, False, False
							print(f'[SERCOTEC] Error: {e}\n{format_exc()}')
				except TimeoutError:
					print('[SERCOTEC] No results')
				except Exception as e:
					if not askokcancel('Web Scraper - ECIT', f'[SERCOTEC] Error: {e}\n{format_exc()}', icon=ERROR):
						return False, False, False
					print(f'[SERCOTEC] Error: {e}\n{format_exc()}')
			browser.close()
	except Exception as e:
		if not askokcancel('Web Scraper - ECIT', f'[SERCOTEC] Error: {e}\n{format_exc()}', icon=ERROR):
			return False, False, False
		print(f'[SERCOTEC] Error: {e}\n{format_exc()}')
	finally:
		return results, count, excluded

'''
┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓
┃				MAIN				┃
┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛'''

def main(ticket_mercado_publico=_DATA['ticket'], keywords=_DATA['keywords'], root=None, log_var = None):
	output = ''
	resultados = []

	if len(ticket_mercado_publico) > 0:
		tmp, x, n = mercado_publico(ticket_mercado_publico, keywords)
		if tmp == False:
			return output
		resultados += tmp
		print(f'Listo Mercado Publico!\n')
		output += f'Mercado Publico:\n{x} resultados incluidos\n{n} resultados excluidos.'
		df = DataFrame(resultados)
		export_to_excel(_DATA['save_path'], df)
		if root and log_var:
			log_var.set(output)
			root.update()

	tmp, x, n = bid()
	if tmp == False:
		return output
	resultados += tmp
	print('listo BID!\n')
	output += f'\n\nBID:\n{x} resultados incluidos\n{n} resultados excluidos.'
	df = DataFrame(resultados)
	export_to_excel(_DATA['save_path'], df)
	if root and log_var:
		log_var.set(output)
		root.update()
	
	tmp, x, n = anid()
	if tmp == False:
		return output
	resultados += tmp
	print('listo ANID!\n')
	output += f'\n\nANID!\n{x} resultados incluidos\n{n} resultados excluidos.'
	df = DataFrame(resultados)
	export_to_excel(_DATA['save_path'], df)
	log_var.set(output)
	if root and log_var:
		log_var.set(output)
		root.update()

	tmp, x, n = sercotec()
	if tmp == False:
		return output
	resultados += tmp
	print('listo SERCOTEC!\n')
	output += f'\n\nSERCOTEC!\n{x} resultados incluidos\n{n} resultados excluidos.'
	df = DataFrame(resultados)
	export_to_excel(_DATA['save_path'], df)
	log_var.set(output)
	if root and log_var:
		log_var.set(output)
		root.update()
	

	x_corfo = 0
	x_codesser = 0
	n_corfo = 0
	n_codesser = 0
	for keyword in keywords:
		tmp, x, n = corfo(keyword)
		if tmp == False:
			return output
		resultados += tmp
		x_corfo += x
		n_corfo += n

		tmp, x, n = codesser(keyword)
		if tmp == False:
			return output
		resultados += tmp
		x_codesser += x
		n_codesser += n
		df = DataFrame(resultados)
		export_to_excel(_DATA['save_path'], df)
		sleep(1)
	output += f'\n\nCORFO:\n{x_corfo} resultados incluidos!\n{n_corfo} resultados excluidos.'
	output += f'\n\nCODESSER:\n{x_codesser} resultados incluidos!\n{n_codesser} resultados excluidos.'
	output += f'\n\nScraping finalizado exitosamente!'
	if root and log_var:
		log_var.set(output)
		root.update()

	df = DataFrame(resultados)
	export_to_excel(_DATA['save_path'], df)
	return output, resultados

if __name__ == '__main__':
	if _DEBUG:
		resultados, x, n = anid()
		print('listo ANID!\n')
		print(f'\n\nANID!\n{x} resultados incluidos\n{n} resultados excluidos.')
		print(resultados)
		df = DataFrame(resultados)
		export_to_excel(_DATA['save_path'], df)
	else:
		main()